
from openpyxl import load_workbook


def update_student():

    # Open Excel file
    wb = load_workbook(
        "students.xlsx"
    )

    sheet = wb.active

    print("\n===== UPDATE STUDENT =====")

    # Take student name
    search_name = input("Enter Student Name: ").strip().lower()

    found = False

    # Loop through rows
    for row in range(2, sheet.max_row + 1):

        current_name = str(sheet.cell(row, 3).value).strip().lower()

        # Match student name
        if current_name == search_name:

            found = True

            # Store old values
            old_name = sheet.cell(row, 3).value
            old_class = sheet.cell(row, 4).value
            old_father = sheet.cell(row, 5).value
            old_mother = sheet.cell(row, 6).value
            old_city = sheet.cell(row, 7).value
            old_phone = sheet.cell(row, 8).value
            old_school = sheet.cell(row, 9).value
            old_status = sheet.cell(row, 10).value
            old_remarks = sheet.cell(row, 11).value

            print("\n===== CURRENT DETAILS =====\n")

            print("Name            :", old_name)
            print("Class           :", old_class)
            print("Father Name     :", old_father)
            print("Mother Name     :", old_mother)
            print("City            :", old_city)
            print("Phone           :", old_phone)
            print("Previous School :", old_school)
            print("Status          :", old_status)
            print("Remarks         :", old_remarks)

            print("\nPress ENTER to keep old value")

            # Take updated values
            new_name = input(f"New Name [{old_name}] : ").strip()
            new_class = input(f"New Class [{old_class}] : ").strip()
            new_father = input(f"New Father Name [{old_father}] : ").strip()
            new_mother = input(f"New Mother Name [{old_mother}] : ").strip()
            new_city = input(f"New City [{old_city}] : ").strip()
            new_phone = input(f"New Phone [{old_phone}] : ").strip()
            new_school = input(f"New Previous School [{old_school}] : ").strip()
            new_status = input(f"New Status [{old_status}] : ").strip()
            new_remarks = input(f"New Remarks [{old_remarks}] : ").strip()

            # Keep old value if input is empty
            sheet.cell(row, 3).value = new_name if new_name else old_name
            sheet.cell(row, 4).value = new_class if new_class else old_class
            sheet.cell(row, 5).value = new_father if new_father else old_father
            sheet.cell(row, 6).value = new_mother if new_mother else old_mother
            sheet.cell(row, 7).value = new_city if new_city else old_city
            sheet.cell(row, 8).value = new_phone if new_phone else old_phone
            sheet.cell(row, 9).value = new_school if new_school else old_school
            sheet.cell(row, 10).value = new_status if new_status else old_status
            sheet.cell(row, 11).value = new_remarks if new_remarks else old_remarks

            # Save Excel file
            wb.save("students.xlsx")

            print("\nStudent Updated Successfully")

            break

    # If student not found
    if found == False:
        print("\nStudent Not Found")

