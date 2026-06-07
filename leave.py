
from openpyxl import load_workbook


def leave_student():

    # Open Excel file
    wb = load_workbook(
        "students.xlsx"
    )

    sheet = wb.active

    print("\n===== STUDENT LEAVE SYSTEM =====")

    # Take student name
    search_name = input("Enter Student Name: ").strip().lower()

    found = False

    # Loop through rows
    for row in range(2, sheet.max_row + 1):

        current_name = str(sheet.cell(row, 3).value).strip().lower()

        # Match student name
        if current_name == search_name:

            found = True

            print("\n===== STUDENT DETAILS =====\n")

            print("Name            :", sheet.cell(row, 3).value)
            print("Class           :", sheet.cell(row, 4).value)
            print("Father Name     :", sheet.cell(row, 5).value)
            print("Phone           :", sheet.cell(row, 8).value)
            print("Current Status  :", sheet.cell(row, 10).value)

            # Check already left
            if str(sheet.cell(row, 10).value).upper() == "LEFT":

                print("\nStudent already left school")
                break

            print("\n===== LEAVE DETAILS =====")

            leave_reason = input("Enter Leave Reason: ").strip()

            leave_remarks = input("Enter Remarks: ").strip()

            # Update status and remarks
            sheet.cell(row, 10).value = "LEFT"

            sheet.cell(row, 11).value = (
                f"LEFT STUDENT | Reason: {leave_reason} | Remarks: {leave_remarks}"
            )

            # Save workbook
            wb.save(
                "students.xlsx"
            )

            print("\nStudent Marked As LEFT Successfully")

            break

    # Student not found
    if found == False:
        print("\nStudent Not Found")

