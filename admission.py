
from openpyxl import load_workbook


def add_student():

    # Open Excel file
    wb = load_workbook("students.xlsx")

    # Select active sheet
    sheet = wb.active

    print("\n--- New Admission ---")

    # Taking student details
    date = input("Enter Date: ")
    name = input("Enter Student Name: ")
    student_class = input("Enter Class: ")
    father_name = input("Enter Father Name: ")
    mother_name = input("Enter Mother Name: ")
    city = input("Enter City: ")
    phone = input("Enter Phone Number: ")
    previous_school = input("Enter Previous School: ")
    remarks = input("Enter Remarks: ")

    # Find next empty row
    row = sheet.max_row + 1

    # Auto generate student ID
    student_id = row - 1

    # Insert data into Excel
    sheet.cell(row, 1).value = student_id
    sheet.cell(row, 2).value = date
    sheet.cell(row, 3).value = name
    sheet.cell(row, 4).value = student_class
    sheet.cell(row, 5).value = father_name
    sheet.cell(row, 6).value = mother_name
    sheet.cell(row, 7).value = city
    sheet.cell(row, 8).value = phone
    sheet.cell(row, 9).value = previous_school
    sheet.cell(row, 10).value = "ACTIVE"
    sheet.cell(row, 11).value = remarks
    # Save Excel file
    wb.save("students.xlsx")

    print("\nStudent Added Successfully")

