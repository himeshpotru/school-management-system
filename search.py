
from openpyxl import load_workbook


def search_student():

    # Open Excel file
    wb = load_workbook("students.xlsx")

    # Select sheet
    sheet = wb.active

    print("\n===== SEARCH STUDENT =====")

    print("1. Search by Student Name")
    print("2. Search by Father Name")
    print("3. Search by Phone Number")

    choice = input("\nEnter Choice: ")

    search_value = input("Enter Search Value: ").lower()

    found = False

    # Loop through rows
    for row in range(2, sheet.max_row + 1):

        student_name = str(sheet.cell(row, 3).value).lower()
        father_name = str(sheet.cell(row, 5).value).lower()
        phone = str(sheet.cell(row, 8).value).lower()

        # Search conditions
        if (
            (choice == "1" and student_name == search_value)
            or
            (choice == "2" and father_name == search_value)
            or
            (choice == "3" and phone == search_value)
        ):

            found = True

            print("\n===== STUDENT FOUND =====\n")

            print("ID               :", sheet.cell(row, 1).value)
            print("Date             :", sheet.cell(row, 2).value)
            print("Name             :", sheet.cell(row, 3).value)
            print("Class            :", sheet.cell(row, 4).value)
            print("Father Name      :", sheet.cell(row, 5).value)
            print("Mother Name      :", sheet.cell(row, 6).value)
            print("City             :", sheet.cell(row, 7).value)
            print("Phone            :", sheet.cell(row, 8).value)
            print("Previous School  :", sheet.cell(row, 9).value)
            print("Status           :", sheet.cell(row, 10).value)
            print("Remarks          :", sheet.cell(row, 11).value)

            break

    if found == False:
        print("\nStudent Not Found")

